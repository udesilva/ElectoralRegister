import argparse
import datetime
import os
import logging
import openpyxl

# Expands the electoral register columns in an unencrypted xlsx spreadsheet.

# Set up logging to file and console
logging.basicConfig(filename='expand-electoral-register.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')
console = logging.StreamHandler()
console.setLevel(logging.INFO)
logging.getLogger('').addHandler(console)

# Create an argument parser
parser = argparse.ArgumentParser(description='Process an unencrypted xlsx spreadsheet.')
parser.add_argument('xlsx_path', type=str,
                    help='Path to the unencrypted xlsx spreadsheet')

# Parse the command line arguments
args = parser.parse_args()

# record the start time
start_time = datetime.datetime.now()

logging.info(f"Reading {args.xlsx_path}")

# Load the decrypted xlsx spreadsheet
with open(args.xlsx_path, 'rb') as f:
    workbook = openpyxl.load_workbook(args.xlsx_path)

# Create a new workbook to store the results
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

logging.info("Processing the spreadsheet")

# Iterate over the rows in the first sheet of the encrypted xlsx spreadsheet
sheet = workbook.active
header = None
for row in sheet.iter_rows():
    if header is None:
        header = row
        # create index by using the values of the cells in the header row to 
        # give the column number by name
        index = {cell.value: i for i, cell in enumerate(header)}
        _markers = index['Elector Markers']
        _postcode = index['PostCode']
        _lastrow = len(header) - 1
        # Copy the header row to the new sheet
        new_sheet.append([cell.value for cell in row])
        continue

    row_val = [str(i.value).replace('\xa0', ' ')
               if i.value is not None else '' for i in row]
    changed = False

    # Exclude "B" and "G" Elector markers for EU citizens who cannot vote in 
    # UK General Elections
    if row_val[_markers] in ["B", "G"]:
        continue

    # Get the postcode fields
    postcode = row_val[_postcode]
    if postcode == '':
        continue

    if postcode in row_val[_postcode+1:]:
        while row_val[_lastrow] != postcode:
            row_val.insert(_postcode+1, '')

        row_val = row_val[:_lastrow+1]
        changed = True

    # Copy the row to the new sheet
    new_sheet.append(row_val)

# Save the results to a new xlsx file

# Get the original filename
original_filename = os.path.basename(args.xlsx_path)

# Add _filtered to the basename
new_filename = os.path.splitext(original_filename)[0] + '_filtered.xlsx'

logging.info(f"Saving to {new_filename}")

# Save the results to the new xlsx file
new_workbook.save(new_filename)

logging.info("Done.")

# record the end time
end_time = datetime.datetime.now()

# Log the time taken
time_diff_seconds = (end_time - start_time).total_seconds()
rounded_seconds = round(time_diff_seconds, 1)

logging.info(f"Time taken: {rounded_seconds} seconds")
